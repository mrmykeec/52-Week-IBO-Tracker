"""Generates tracker.xlsx — the 52-week IBO tracker for Google Sheets."""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
import os

# ── DATA ──────────────────────────────────────────────────────────────────────

MONTHLY_GOALS = {
    1:  ["150pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "1+ New Customer",
         "4+ New Prospects added to your list",
         "Showed 1+ Plan (with mentor, on your own, or Board Plan / WWG event)",
         "Listened to 8+ Audios",
         "Read 1+ Book — Recommendation: Simply Rich by Rich DeVos",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Scheduled Retail Event for next month",
         "Sent 4+ Topics to Coach every week with audio takeaways or questions"],
    2:  ["150pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "1+ New Customer",
         "5+ New Prospects added to your list",
         "Add 1+ new product to your DITTO",
         "Showed 2+ Plans",
         "Listened to 8+ Audios",
         "Read 1+ Book — Recommendation: True North by Georgia Lee Puryear",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Scheduled Health Break for Week 10",
         "Sent 5+ Topics to Coach every week with audio takeaways or questions"],
    3:  ["150pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "1+ New Customer",
         "6+ New Prospects added to your list",
         "Add 1+ new product to your DITTO",
         "Showed 2+ Plans",
         "Listened to 8+ Audios",
         "Read 1+ Book — Recommendation: Magic of Thinking Big by David Schwartz",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Completed Health Break (Week 10)",
         "Sent 6+ Topics to Coach every week with audio takeaways or questions"],
    4:  ["200pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "2+ New Customers",
         "6+ New Prospects added to your list",
         "Add 1+ new product to your DITTO",
         "Showed 3+ Plans",
         "Listened to 8+ Audios",
         "Read 1+ Book — Recommendation: Developing the Leader Within You by John Maxwell",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Created Customer Binder",
         "Sent 6+ Topics to Coach every week with audio takeaways or questions"],
    5:  ["200pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "2+ New Customers",
         "8+ New Prospects added to your list",
         "Add 1+ new product to your DITTO",
         "Showed 3+ Plans",
         "Listened to 8+ Audios",
         "Read 1+ Book — Recommendation: Cashflow Quadrant by Robert Kiyosaki",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Scheduled Beauty Break for Week 21",
         "Sent 7+ Topics to Coach every week with audio takeaways or questions"],
    6:  ["200pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "2+ New Customers",
         "8+ New Prospects added to your list",
         "Add 1+ new product to your DITTO",
         "Showed 4+ Plans",
         "Listened to 8+ Audios",
         "Read 1+ Book",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Completed Beauty Break (Week 21)",
         "Sent 8+ Topics to Coach every week with audio takeaways or questions"],
    7:  ["200pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "2+ New Customers",
         "10+ New Prospects added to your list",
         "Add 1+ new product to your DITTO",
         "Showed 4+ Plans",
         "Listened to 8+ Audios",
         "Read 1+ Book",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Scheduled Begin 30 Event for Week 32",
         "Sent 8+ Topics to Coach every week with audio takeaways or questions"],
    8:  ["250pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "3+ New Customers",
         "10+ New Prospects added to your list",
         "Add 1+ new product to your DITTO",
         "Showed 5+ Plans",
         "Listened to 8+ Audios",
         "Read 1+ Book",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Completed Begin 30 Event (Week 32)",
         "Sent 10+ Topics to Coach every week with audio takeaways or questions"],
    9:  ["250pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "3+ New Customers",
         "12+ New Prospects added to your list",
         "Add 1+ new product to your DITTO",
         "Showed 5+ Plans",
         "Listened to 8+ Audios",
         "Read 1+ Book",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Sent 12+ Topics to Coach every week with audio takeaways or questions"],
    10: ["250pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "3+ New Customers",
         "12+ New Prospects added to your list",
         "Add 1+ new product to your DITTO",
         "Showed 6+ Plans",
         "Listened to 8+ Audios",
         "Read 1+ Book",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Achieved a Qualified Month",
         "Sent 14+ Topics to Coach every week with audio takeaways or questions"],
    11: ["250pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "3+ New Customers",
         "15+ New Prospects added to your list",
         "Add 1+ new product to your DITTO",
         "Showed 6+ Plans",
         "Listened to 8+ Audios",
         "Read 1+ Book",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Created Vision / Dream Board",
         "Sent 14+ Topics to Coach every week with audio takeaways or questions"],
    12: ["300pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "3+ New Customers",
         "15+ New Prospects added to your list",
         "Add 1+ new product to your DITTO",
         "Showed 8+ Plans",
         "Listened to 8+ Audios",
         "Read 1+ Book",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Scheduled Amway Home Retail Event for Week 50",
         "Sent 16+ Topics to Coach every week with audio takeaways or questions"],
    13: ["300pv Minimum Personal Circle / 60% Verified Customer Sales (VCS)",
         "3+ New Customers",
         "15+ New Prospects added to your list",
         "Add 1+ new product to your DITTO",
         "Showed 8+ Plans",
         "Listened to 8+ Audios",
         "Read 1+ Book",
         "Attended Team Meeting, WWG Event, 2+ Board Plans, Howie Livestream",
         "Completed Amway Home Retail Event (Week 50)",
         "Sent 16+ Topics to Coach every week",
         "GREAT JOB! You finished the 52-Week Training. Now… do it again!"],
}

WEEKS = [
    (1,1,"learning",'Listen: "Kenny and Ashlea Toms Diamond Seminar" — WWG App / wwdb.com > WWG Store > Search "1237"'),
    (1,1,"learning",'Watch: "Activating New IBOs" — championleadership1.com > Training Videos (Password: vision3)'),
    (1,1,"learning",'Watch: "Basics of the Business" — championleadership1.com > Training Videos (Password: vision3)'),
    (1,1,"event",   "Place Amway Order — include Starter Stacks (first 30 days). Search 'Starter Stacks' in Amway App or amway.com"),
    (1,1,"event",   "Create Your Make Aware — your personal story of how you found this industry. Send Topic to your coach with your story"),
    (1,1,"event",   "Get added to Howie and Theresa's Topics and monthly Newsletter — ask your coach for the Topics numbers"),
    (1,1,"habit",   "Clear Topics Daily"),
    (1,1,"habit",   "Check DreamStream Daily"),
    (1,1,"habit",   "Send Topic to coach with your takeaways and a question"),

    (2,1,"learning",'Listen: "Las Vegas Retail Panel 1, 2, 3, 4" — WWG App > WWG Store > Search "RTLPNL"'),
    (2,1,"learning","Watch: Retail Shelf IBO Training — WWG App > Retail shelf > 'For IBOs' icon > Study Steps 1-6"),
    (2,1,"learning",'Understand $300 VCS Requirement, 60% VCS, CSI, and SSI — Reference "Money and Rewards Brochure" (WWG App > Info Center > PDFs > "Money" > Pages 3, 5-6)'),
    (2,1,"event",   "Save your MyShop retail website to your phone home screen"),
    (2,1,"networking","Register a customer using your MyShop retail website"),
    (2,1,"event",   "Create a receipt for a customer sale"),
    (2,1,"learning","Understand the upcoming contest (Diamond Day / Family Fun Day) — visit Danzik Motivation Newsletter for the flyer"),
    (2,1,"networking","Practice your Make Aware with someone (friend, family, or stranger) and coach up on it"),
    (2,1,"habit",   "Clear Topics Daily"),
    (2,1,"habit",   "Check DreamStream Daily"),
    (2,1,"habit",   "Send Topic to coach with reflection on your Make Aware, takeaways, or questions"),

    (3,1,"learning",'Listen: "Howie Danzik Wrap-Up" — WWG App > WWG Store > Search "1376"'),
    (3,1,"learning",'Watch: "Retailing Tips" — championleadership1.com > Training Videos (Password: vision3)'),
    (3,1,"learning",'Watch: "Customer Sales Incentive (CSI)" — Amway App / amway.com > Education > Search "Customer Sales Incentive"'),
    (3,1,"event",   'Create your Client and Prospect List — Reference "Building Your Name List" (WWG App > Info Center > PDFs > "Building")'),
    (3,1,"event",   "Schedule a Retail Event (virtual or live, with mentor or on your own, one-on-one or group)"),
    (3,1,"habit",   "Clear Topics Daily"),
    (3,1,"habit",   "Check DreamStream Daily"),
    (3,1,"habit",   "Send Topic to coach with takeaways and a question"),

    (4,1,"learning",'Listen: "Spring Leadership 2013 — Build Your Width" — WWG App > WWG Store > Search "724"'),
    (4,1,"learning",'Watch: "What Differentiates Amway" — championleadership1.com > Training Videos'),
    (4,1,"learning","Learn about DITTO (auto-reorder) — watch the DITTO video and reference the DITTO PDFs in the WWG App"),
    (4,1,"event",   "Set up your DITTO (auto-reorder) with at least one Amway product"),
    (4,1,"networking","Practice Interjections: say 'Hi' to everyone you encounter this week"),
    (4,1,"habit",   "Clear Topics Daily"),
    (4,1,"habit",   "Check DreamStream Daily"),
    (4,1,"habit",   "Send Topic to coach with takeaways and a question"),

    (5,2,"learning",'Listen: "Howie & Theresa Seminar" — WWG App > WWG Store > Search "702"'),
    (5,2,"learning",'Watch: "People Skills" — championleadership1.com > Training Videos'),
    (5,2,"networking","Practice Interjections: give a compliment + ask a question to 1+ person per day"),
    (5,2,"networking","Call 3 people to share your business or retail website"),
    (5,2,"habit",   "Clear Topics Daily"),
    (5,2,"habit",   "Check DreamStream Daily"),
    (5,2,"habit",   "Send Topic to coach with takeaways and a question"),

    (6,2,"learning",'Listen: "Try or Cry" — WWG App > WWG Store > Search "AM108"'),
    (6,2,"learning",'Watch: "How to Get a Customer Order" — championleadership1.com > Training Videos'),
    (6,2,"learning",'Watch: "SSI" — Amway App / amway.com > Education > Search "SSI"'),
    (6,2,"networking","Practice Interjections: say 'Hi, how are you?' to everyone you encounter"),
    (6,2,"habit",   "Clear Topics Daily"),
    (6,2,"habit",   "Check DreamStream Daily"),
    (6,2,"habit",   "Send Topic to coach with takeaways and a question"),

    (7,2,"learning",'Listen: "SL 2023 — Toms and Erickson" — WWG App > WWG Store > Search "1367"'),
    (7,2,"learning",'Watch: "Power of Being Core" — championleadership1.com > Training Videos'),
    (7,2,"learning",'Print and study: "Core: 10 Habits" and "Core Run Monthly Worksheet" — WWG App > Info Center > PDFs'),
    (7,2,"networking","Practice Interjections: give a genuine compliment to everyone you meet"),
    (7,2,"event",   "Promote your Week 10 Retail Event to potential guests"),
    (7,2,"habit",   "Clear Topics Daily"),
    (7,2,"habit",   "Check DreamStream Daily"),
    (7,2,"habit",   "Send Topic to coach with takeaways and a question"),

    (8,2,"learning",'Listen: "FED 2022 — Brown and Ericksons" — WWG App > WWG Store > Search "1345"'),
    (8,2,"learning",'Watch: "Maximizing Your Volume" — championleadership1.com > Training Videos'),
    (8,2,"learning",'Watch: "Starting Conversations Naturally" — Amway App / amway.com > Education'),
    (8,2,"networking","Practice Interjections: 5 Make Awares using your personal story this week"),
    (8,2,"habit",   "Clear Topics Daily"),
    (8,2,"habit",   "Check DreamStream Daily"),
    (8,2,"habit",   "Send Topic to coach with takeaways and a question"),

    (9,3,"learning",'Listen: "FED 2021 — Jassman and Whalen" — WWG App > WWG Store > Search "1288"'),
    (9,3,"learning",'Watch: "Retailing Events" — championleadership1.com > Training Videos'),
    (9,3,"learning",'Watch: "Personal Volume" — Amway App / amway.com > Education'),
    (9,3,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (9,3,"habit",   "Clear Topics Daily"),
    (9,3,"habit",   "Check DreamStream Daily"),
    (9,3,"habit",   "Send Topic to coach with takeaways and a question"),

    (10,3,"learning",'Listen: "Greg Duncan — Network Marketing" — WWG App > WWG Store > Search "1303"'),
    (10,3,"learning",'Watch: "How to Do a Health Break" — championleadership1.com > Training Videos'),
    (10,3,"learning","Study Health Break Outline — Danzik Motivation Newsletter"),
    (10,3,"event",  "Complete a Health Break with 1+ person (virtual or live)"),
    (10,3,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (10,3,"habit",  "Clear Topics Daily"),
    (10,3,"habit",  "Check DreamStream Daily"),
    (10,3,"habit",  "Send Topic to coach with takeaways and a question"),

    (11,3,"learning",'Listen: "SL 2022 — Ewing and Yuen" — WWG App > WWG Store > Search "1315"'),
    (11,3,"learning",'Watch: "PV/BV" — Amway App / amway.com > Education'),
    (11,3,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (11,3,"networking","Follow up with Health Break customers — check if they received their orders"),
    (11,3,"networking","Interjection outing: mall or park — 2 Make Awares + 1 Drop Message"),
    (11,3,"habit",  "Clear Topics Daily"),
    (11,3,"habit",  "Check DreamStream Daily"),
    (11,3,"habit",  "Send Topic to coach with takeaways and a question"),

    (12,3,"learning",'Listen: "FED 2021 — Olynyk and Eaton" — WWG App > WWG Store > Search "1281"'),
    (12,3,"learning",'Watch: "3 Things You Must Know to Get Started in Retail" — championleadership1.com'),
    (12,3,"learning",'Watch: "Personal Performance Bonus" — Amway App / amway.com > Education'),
    (12,3,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (12,3,"networking","Follow up with Health Break customers — check if they are enjoying their orders"),
    (12,3,"habit",  "Clear Topics Daily"),
    (12,3,"habit",  "Check DreamStream Daily"),
    (12,3,"habit",  "Send Topic to coach with takeaways and a question"),

    (13,4,"learning",'Listen: "Carrithers Emerald Rally" — WWG App > WWG Store > Search "1323"'),
    (13,4,"learning","Listen: Money Management Conference Call — WWG App > Topics"),
    (13,4,"learning",'Watch: "Money Management" — championleadership1.com > Training Videos'),
    (13,4,"event",  "Complete a Budget — print and fill out the Budget Worksheet (WWG App > Info Center > PDFs)"),
    (13,4,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (13,4,"habit",  "Clear Topics Daily"),
    (13,4,"habit",  "Check DreamStream Daily"),
    (13,4,"habit",  "Send Topic to coach with takeaways and a question"),

    (14,4,"learning",'Listen: "SL 2022 — Whalen and Hawkins" — WWG App > WWG Store > Search "1310"'),
    (14,4,"learning",'Watch: "Organizing Your Customers" — championleadership1.com > Training Videos'),
    (14,4,"learning","Review Customer Profile Form and Quick Customer Survey — WWG App > Info Center > PDFs"),
    (14,4,"event",  "Create your Customer Binder — organize existing customers with their profiles"),
    (14,4,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (14,4,"habit",  "Clear Topics Daily"),
    (14,4,"habit",  "Check DreamStream Daily"),
    (14,4,"habit",  "Send Topic to coach with takeaways and a question"),

    (15,4,"learning",'Listen: "Diamond Seminar — Olynyk" — WWG App > WWG Store > Search "1313"'),
    (15,4,"learning",'Watch: "Dress for Success" — championleadership1.com > Training Videos'),
    (15,4,"learning",'Watch: "Group Volume" — Amway App / amway.com > Education'),
    (15,4,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (15,4,"networking","Interjection outing: gym or Farmer's Market — 2 Make Awares + 1 Drop Message"),
    (15,4,"habit",  "Clear Topics Daily"),
    (15,4,"habit",  "Check DreamStream Daily"),
    (15,4,"habit",  "Send Topic to coach with takeaways and a question"),

    (16,4,"learning",'Listen: "FED 2022 — Rosario and Rice" — WWG App > WWG Store > Search "1343"'),
    (16,4,"learning",'Watch: "Business Structure" — Amway App / amway.com > Education'),
    (16,4,"learning",'Study "Depth vs. Width" — Business Gameplan PDF (WWG App > Info Center > PDFs)'),
    (16,4,"learning","Study Eagle and Double Eagle qualifications — PDF in WWG App > Info Center"),
    (16,4,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (16,4,"habit",  "Clear Topics Daily"),
    (16,4,"habit",  "Check DreamStream Daily"),
    (16,4,"habit",  "Send Topic to coach with takeaways and a question"),

    (17,5,"learning",'Listen: "FED 2021 — Phillips, Ewing, Yadao" — WWG App > WWG Store > Search "1280"'),
    (17,5,"learning",'Watch: "Use MyShop to Build Your Business" — Amway App / amway.com > Education'),
    (17,5,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (17,5,"habit",  "Clear Topics Daily"),
    (17,5,"habit",  "Check DreamStream Daily"),
    (17,5,"habit",  "Send Topic to coach with takeaways and a question"),

    (18,5,"learning",'Listen: "SL Diamonds Men\'s Panel — St. Paul 2023" — WWG App > WWG Store > Search "1365"'),
    (18,5,"learning",'Watch: "Cashflow Quadrant" — championleadership1.com > Training Videos'),
    (18,5,"learning","Learn about Amway — weareamway.com, iboai.com, ibofacts.com"),
    (18,5,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (18,5,"networking","Find a local event — get 2 Make Awares + 1 Drop Message"),
    (18,5,"habit",  "Clear Topics Daily"),
    (18,5,"habit",  "Check DreamStream Daily"),
    (18,5,"habit",  "Send Topic to coach with takeaways and a question"),

    (19,5,"learning",'Listen: "FED 2022 — Olynyk" — WWG App > WWG Store > Search "1350"'),
    (19,5,"learning","Listen: Mentorship Conference Call — WWG App > Topics"),
    (19,5,"learning",'Watch: "Differential Bonus" — Amway App / amway.com > Education'),
    (19,5,"event",  "Promote your Week 21 Retail Event (Beauty Break) to potential guests"),
    (19,5,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (19,5,"habit",  "Clear Topics Daily"),
    (19,5,"habit",  "Check DreamStream Daily"),
    (19,5,"habit",  "Send Topic to coach with takeaways and a question"),

    (20,5,"learning",'Listen: "SL 2023 — Herschelman and Felber" — WWG App > WWG Store > Search "1368"'),
    (20,5,"learning",'Watch: "Simplified Beauty Break" — championleadership1.com > Training Videos'),
    (20,5,"learning","Learn about Artistry — WWG App > Retail > Beauty > 2026 Artistry Brand Video"),
    (20,5,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (20,5,"habit",  "Clear Topics Daily"),
    (20,5,"habit",  "Check DreamStream Daily"),
    (20,5,"habit",  "Send Topic to coach with takeaways and a question"),

    (21,6,"learning",'Listen: "Now is the Time" — WWG App > WWG Store > Search "AM104"'),
    (21,6,"learning","Learn Artistry Virtual Beauty Tool — amway.com > Artistry > Virtual Try-On"),
    (21,6,"event",  "Complete a virtual or live Beauty Break with 1+ person"),
    (21,6,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (21,6,"habit",  "Clear Topics Daily"),
    (21,6,"habit",  "Check DreamStream Daily"),
    (21,6,"habit",  "Send Topic to coach with takeaways and a question"),

    (22,6,"learning",'Listen: "SL 2023 — Tuitupou and Hawkins" — WWG App > WWG Store > Search "1381"'),
    (22,6,"learning",'Watch: "Reasons for Big Business" — championleadership1.com > Training Videos'),
    (22,6,"learning",'Watch: "Differential Bonus with Multiple Frontlines" — Amway App / amway.com > Education'),
    (22,6,"networking","Practice Interjections: 5 Make Awares + 2 Drop Messages this week"),
    (22,6,"networking","Attend a networking event — get 1 Make Aware + 1 Drop Message"),
    (22,6,"networking","Follow up with Beauty Break customers — check if they received their orders"),
    (22,6,"habit",  "Clear Topics Daily"),
    (22,6,"habit",  "Check DreamStream Daily"),
    (22,6,"habit",  "Send Topic to coach with takeaways and a question"),

    (23,6,"learning",'Listen: "Diamond Bio — Kummer" — WWG App > WWG Store > Search "DB433"'),
    (23,6,"learning","Listen: FAQs Conference Call — WWG App > Topics"),
    (23,6,"learning",'Watch: "Sell eSpring Water Purifier — Start Selling" — Amway App / amway.com > Education'),
    (23,6,"networking","Practice Interjections: 7 Make Awares + 3 Drop Messages this week"),
    (23,6,"networking","Follow up with Beauty Break customers — check if they are enjoying their orders"),
    (23,6,"networking","Talk to one person about eSpring Water Purifier"),
    (23,6,"habit",  "Clear Topics Daily"),
    (23,6,"habit",  "Check DreamStream Daily"),
    (23,6,"habit",  "Send Topic to coach with takeaways and a question"),

    (24,6,"learning",'Listen: "Family Reunion 2023 — Rosario and Eaton" — WWG App > WWG Store > Search "1396"'),
    (24,6,"learning",'Watch: "Customer Tips and Mindset" — championleadership1.com > Training Videos'),
    (24,6,"learning",'Watch: "Sell Atmosphere Sky and Mini Air: Start Selling" — Amway App / amway.com > Education'),
    (24,6,"networking","Practice Interjections: 7 Make Awares + 3 Drop Messages this week"),
    (24,6,"habit",  "Clear Topics Daily"),
    (24,6,"habit",  "Check DreamStream Daily"),
    (24,6,"habit",  "Send Topic to coach with takeaways and a question"),

    (25,7,"learning",'Listen: "FED 2023 — Herschelman" — WWG App > WWG Store > Search "1403"'),
    (25,7,"learning",'Watch: "Things I\'ve Learned from Mentors" — championleadership1.com > Training Videos'),
    (25,7,"learning",'Watch: "Sell Atmosphere Sky: Price Value" — Amway App / amway.com > Education'),
    (25,7,"networking","Practice Interjections: 7 Make Awares + 3 Drop Messages this week"),
    (25,7,"networking","Talk to one person about Atmosphere Air Treatment products"),
    (25,7,"habit",  "Clear Topics Daily"),
    (25,7,"habit",  "Check DreamStream Daily"),
    (25,7,"habit",  "Send Topic to coach with takeaways and a question"),

    (26,7,"learning",'Listen: "Diamond Seminar — Tuitupou" — WWG App > WWG Store > Search "1434"'),
    (26,7,"learning",'Watch: "Bronze Foundation Incentive (BFI)" — Amway App / amway.com > Education'),
    (26,7,"learning","Learn about Bronze Foundation — Core Plus Incentives 2025 PDF (WWG App > Info Center > PDFs)"),
    (26,7,"networking","Practice Interjections: 7 Make Awares + 3 Drop Messages this week"),
    (26,7,"networking","Find a local event — get 2 Make Awares + 1 Drop Message"),
    (26,7,"habit",  "Clear Topics Daily"),
    (26,7,"habit",  "Check DreamStream Daily"),
    (26,7,"habit",  "Send Topic to coach with takeaways and a question"),

    (27,7,"learning",'Listen: "SL 2023 — Rosario and Kummer" — WWG App > WWG Store > Search "1384"'),
    (27,7,"learning","Listen: Handling Objections Conference Call — WWG App > Topics"),
    (27,7,"learning",'Watch: "Selling Nutrilite Organics" — Amway App / amway.com > Education'),
    (27,7,"networking","Practice Interjections: 7 Make Awares + 3 Drop Messages this week"),
    (27,7,"habit",  "Clear Topics Daily"),
    (27,7,"habit",  "Check DreamStream Daily"),
    (27,7,"habit",  "Send Topic to coach with takeaways and a question"),

    (28,7,"learning",'Listen: "FED 2023 — Kelly and Darci Ewing" — WWG App > WWG Store > Search "1411"'),
    (28,7,"learning",'Watch: "2023 Product Blast" — WWG App > Info Center > PDFs'),
    (28,7,"learning",'Watch: "Nutrilite Featured Products" — Amway App / amway.com > Education'),
    (28,7,"networking","Practice Interjections: 7 Make Awares + 3 Drop Messages this week"),
    (28,7,"habit",  "Clear Topics Daily"),
    (28,7,"habit",  "Check DreamStream Daily"),
    (28,7,"habit",  "Send Topic to coach with takeaways and a question"),

    (29,8,"learning",'Listen: "FED 2023 — Trevor and Lexi Baker" — WWG App > WWG Store > Search "1413"'),
    (29,8,"learning",'Watch: "Things I Love About the Business" — championleadership1.com > Training Videos'),
    (29,8,"learning",'Watch: "Nutrilite Traceability" — Amway App / amway.com > Education'),
    (29,8,"networking","Practice Interjections: 7 Make Awares + 3 Drop Messages this week"),
    (29,8,"habit",  "Clear Topics Daily"),
    (29,8,"habit",  "Check DreamStream Daily"),
    (29,8,"habit",  "Send Topic to coach with takeaways and a question"),

    (30,8,"learning",'Listen: "FED 2023 — Howie and Theresa Danzik" — WWG App > WWG Store > Search "1422"'),
    (30,8,"learning","Listen: Time Management Conference Call — WWG App > Topics"),
    (30,8,"learning",'Watch: "Nutrilite Double X 1: Start Selling" — Amway App / amway.com > Education'),
    (30,8,"networking","Interjection outing: networking event — 3 Make Awares + 2 Drop Messages"),
    (30,8,"networking","Call 3 people to share your MyShop retail website or business opportunity"),
    (30,8,"event",  "Promote your Week 32 Begin 30 Event to potential guests"),
    (30,8,"networking","Talk to 1+ person about Nutrilite Double X"),
    (30,8,"habit",  "Clear Topics Daily"),
    (30,8,"habit",  "Check DreamStream Daily"),
    (30,8,"habit",  "Send Topic to coach with takeaways and a question"),

    (31,8,"learning",'Listen: "Family Reunion 2019 — Micahl and Lorena Wood" — WWG App > WWG Store > Search "1198"'),
    (31,8,"learning","Listen: Scheduling Conference Call + write out your schedule for the upcoming week — WWG App > Topics"),
    (31,8,"learning",'Watch: "Nutrilite Double X 2: PhytoProtect Blend" — Amway App / amway.com > Education'),
    (31,8,"networking","Practice Interjections: 7 Make Awares + 3 Drop Messages this week"),
    (31,8,"habit",  "Clear Topics Daily"),
    (31,8,"habit",  "Check DreamStream Daily"),
    (31,8,"habit",  "Send Topic to coach with takeaways and a question"),

    (32,8,"learning",'Listen: "Freedom" — WWG App > WWG Store > Search "AM105"'),
    (32,8,"learning",'Watch: "Begin 30 Training" — championleadership1.com > Training Videos'),
    (32,8,"learning","Learn the Begin 30 Outline — Danzik Motivation Newsletter"),
    (32,8,"learning","Learn Begin 30 Key Points — WWG App > Info Center"),
    (32,8,"learning",'Watch: "Begin 30 by Ron Tilles" video — WWG App > Info Center'),
    (32,8,"event",  "Complete a virtual or live Begin 30 Event with 1+ person"),
    (32,8,"networking","Practice Interjections: 7 Make Awares + 3 Drop Messages this week"),
    (32,8,"habit",  "Clear Topics Daily"),
    (32,8,"habit",  "Check DreamStream Daily"),
    (32,8,"habit",  "Send Topic to coach with takeaways and a question"),

    (33,9,"learning",'Listen: "Diamond Bio — Bill and Sandy Hawkins" — WWG App > WWG Store > Search "1169"'),
    (33,9,"learning",'Watch: "Bronze Builder Incentive (BBI)" — Amway App / amway.com > Education'),
    (33,9,"learning","Learn about Bronze Builder — Core Plus Incentives 2025 PDF (WWG App > Info Center > PDFs)"),
    (33,9,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (33,9,"networking","Follow up with Begin 30 customers — check if they received their orders"),
    (33,9,"habit",  "Clear Topics Daily"),
    (33,9,"habit",  "Check DreamStream Daily"),
    (33,9,"habit",  "Send Topic to coach with takeaways and a question"),

    (34,9,"learning",'Listen: "Family Reunion 2023 — Kummer and Hawkins" — WWG App > WWG Store > Search "1389"'),
    (34,9,"learning",'Watch: "Follow Up & Follow Through: Serving & Thanking" and "Organizational System" videos'),
    (34,9,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (34,9,"networking","Find a local event — get 3 Make Awares + 2 Drop Messages"),
    (34,9,"networking","Follow up with Begin 30 customers — check if they are enjoying their orders"),
    (34,9,"habit",  "Clear Topics Daily"),
    (34,9,"habit",  "Check DreamStream Daily"),
    (34,9,"habit",  "Send Topic to coach with takeaways and a question"),

    (35,9,"learning",'Listen: "SL 2024 — Jon and Jen Rosario" — WWG App > WWG Store > Search "1448"'),
    (35,9,"learning","Listen: Upline Relationships Conference Call — WWG App > Topics"),
    (35,9,"learning",'Watch: "Nutrilite Double X 3: Price Value" — Amway App / amway.com > Education'),
    (35,9,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (35,9,"habit",  "Clear Topics Daily"),
    (35,9,"habit",  "Check DreamStream Daily"),
    (35,9,"habit",  "Send Topic to coach with takeaways and a question"),

    (36,9,"learning",'Listen: "Diamond Seminar — Shane and Joey Yadao" — WWG App > WWG Store > Search "1209"'),
    (36,9,"learning",'Watch: "FAQs Training" — championleadership1.com > Training Videos'),
    (36,9,"learning",'Watch: "Probiotics: Understand the Basics" — Amway App / amway.com > Education'),
    (36,9,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (36,9,"networking","Interjection outing: walk the mall — 3 Make Awares + 2 Drop Messages"),
    (36,9,"habit",  "Clear Topics Daily"),
    (36,9,"habit",  "Check DreamStream Daily"),
    (36,9,"habit",  "Send Topic to coach with takeaways and a question"),

    (37,10,"learning",'Listen: "SL 2023 — Ewing and Phillips" — WWG App > WWG Store > Search "1366"'),
    (37,10,"learning","Listen: 10 Core Steps Conference Call — WWG App > Topics"),
    (37,10,"learning",'Watch two "Sell Nutrilite Probiotics" videos — Amway App / amway.com > Education'),
    (37,10,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (37,10,"networking","Talk to someone about Nutrilite Probiotics"),
    (37,10,"habit",  "Clear Topics Daily"),
    (37,10,"habit",  "Check DreamStream Daily"),
    (37,10,"habit",  "Send Topic to coach with takeaways and a question"),

    (38,10,"learning",'Listen: "Pump in the Desert" — WWG App > WWG Store > Search "762"'),
    (38,10,"learning",'Watch: "Leadership Observations" — championleadership1.com > Training Videos'),
    (38,10,"learning",'Watch: "Get Started on Social Media" — Amway App / amway.com > Education'),
    (38,10,"learning","Review Social Media Guidelines PDF — WWG App > Info Center > PDFs"),
    (38,10,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (38,10,"habit",  "Clear Topics Daily"),
    (38,10,"habit",  "Check DreamStream Daily"),
    (38,10,"habit",  "Send Topic to coach with takeaways and a question"),

    (39,10,"learning",'Listen: "SL 2024 — Glen and Joya Baker" — WWG App > WWG Store > Search "1462"'),
    (39,10,"learning","Listen: Scheduling Conference Call — write out your schedule for the week — WWG App > Topics"),
    (39,10,"learning",'Watch: "Understand Business Accounts on Social Media" — Amway App / amway.com > Education'),
    (39,10,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (39,10,"networking","Call 3 people to share MyShop retail website or business opportunity"),
    (39,10,"habit",  "Clear Topics Daily"),
    (39,10,"habit",  "Check DreamStream Daily"),
    (39,10,"habit",  "Send Topic to coach with takeaways and a question"),

    (40,10,"learning",'Listen: "SL 2019 — Tracey and Kimberley Eaton" — WWG App > WWG Store > Search "1160"'),
    (40,10,"learning","Listen: Big Business vs. Small Business Conference Call — WWG App > Topics"),
    (40,10,"learning",'Watch: "Achieve a Qualified Month" — Amway App / amway.com > Education'),
    (40,10,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (40,10,"networking","Attend a local event — 3 Make Awares + 2 Drop Messages"),
    (40,10,"habit",  "Clear Topics Daily"),
    (40,10,"habit",  "Check DreamStream Daily"),
    (40,10,"habit",  "Send Topic to coach with takeaways and a question"),

    (41,11,"learning",'Listen: "SL 2024 — Dan and Sandy Yuen" — WWG App > WWG Store > Search "1455"'),
    (41,11,"learning",'Watch: "Leadership Tips" — championleadership1.com > Training Videos'),
    (41,11,"learning",'Watch: "Prepare to Break a Leg" — Amway App / amway.com > Education'),
    (41,11,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (41,11,"networking","Talk to 1+ person about your favorite Amway product"),
    (41,11,"habit",  "Clear Topics Daily"),
    (41,11,"habit",  "Check DreamStream Daily"),
    (41,11,"habit",  "Send Topic to coach with takeaways and a question"),

    (42,11,"learning",'Listen: "Overcoming Offenses" — WWG App > WWG Store > Search "774"'),
    (42,11,"learning","Listen: Helpful Tips Conference Call — WWG App > Topics"),
    (42,11,"learning",'Watch: "Two-Time Cash Incentives (TTCI)" — Amway App / amway.com > Education'),
    (42,11,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (42,11,"networking","Learn about Muscle Multiplier — share with 1+ person"),
    (42,11,"event",  "Create your Vision / Dream Board — images and goals that represent your why"),
    (42,11,"habit",  "Clear Topics Daily"),
    (42,11,"habit",  "Check DreamStream Daily"),
    (42,11,"habit",  "Send Topic to coach with takeaways and a question"),

    (43,11,"learning",'Listen: "SL 2022 — Rosario and Crandell" — WWG App > WWG Store > Search "1319"'),
    (43,11,"learning",'Watch: "Creating a Positive Atmosphere" — Amway App / amway.com > Education'),
    (43,11,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (43,11,"networking","Interjection outing: mall or Farmer's Market — 2 Make Awares + 1 Drop Message"),
    (43,11,"habit",  "Clear Topics Daily"),
    (43,11,"habit",  "Check DreamStream Daily"),
    (43,11,"habit",  "Send Topic to coach with takeaways and a question"),

    (44,11,"learning",'Listen: "10 Points that Make Amway Special" — WWG App > WWG Store > Search "AM106"'),
    (44,11,"learning",'Watch: "Focus on Purpose" — Amway App / amway.com > Education'),
    (44,11,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (44,11,"networking","Learn about LongXevity Eye Cream — tell 1+ person about it"),
    (44,11,"habit",  "Clear Topics Daily"),
    (44,11,"habit",  "Check DreamStream Daily"),
    (44,11,"habit",  "Send Topic to coach with takeaways and a question"),

    (45,12,"learning",'Listen: "FED 2019 — Bob and Shelly Kummer" — WWG App > WWG Store > Search "1204"'),
    (45,12,"learning",'Watch: "The Focus of Success" — Amway App / amway.com > Education'),
    (45,12,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (45,12,"networking","Learn about Nutrilite Energy and Focus supplement — tell 1+ person about it"),
    (45,12,"habit",  "Clear Topics Daily"),
    (45,12,"habit",  "Check DreamStream Daily"),
    (45,12,"habit",  "Send Topic to coach with takeaways and a question"),

    (46,12,"learning",'Listen: "FED 2020 — Howie and Theresa Danzik" — WWG App > WWG Store > Search "1246"'),
    (46,12,"learning",'Watch: "Monthly Leadership Bonus" — Amway App / amway.com > Education'),
    (46,12,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (46,12,"networking","Learn about XS Energy drinks — tell 1+ person about them"),
    (46,12,"habit",  "Clear Topics Daily"),
    (46,12,"habit",  "Check DreamStream Daily"),
    (46,12,"habit",  "Send Topic to coach with takeaways and a question"),

    (47,12,"learning",'Listen: "FED 2019 — Jon and Jen Rosario" — WWG App > WWG Store > Search "1230"'),
    (47,12,"learning",'Watch: "Raise the Main" — Amway App / amway.com > Education'),
    (47,12,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (47,12,"networking","Learn about Amway Home Laundry products — tell 1+ person about them"),
    (47,12,"habit",  "Clear Topics Daily"),
    (47,12,"habit",  "Check DreamStream Daily"),
    (47,12,"habit",  "Send Topic to coach with takeaways and a question"),

    (48,12,"learning",'Listen: "Family Reunion 2019 — Jake and Jolene Carrithers" — WWG App > WWG Store > Search "1189"'),
    (48,12,"learning",'Watch: "Resilience is a Muscle" — Amway App / amway.com > Education'),
    (48,12,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (48,12,"event",  "Promote your Week 50 Amway Home Retail Event to potential guests"),
    (48,12,"networking","Learn about Amway Home surface cleaners — tell 1+ person about them"),
    (48,12,"habit",  "Clear Topics Daily"),
    (48,12,"habit",  "Check DreamStream Daily"),
    (48,12,"habit",  "Send Topic to coach with takeaways and a question"),

    (49,13,"learning",'Listen: "Diamond Bio — Brad and Julie Duncan" — WWG App > WWG Store > Search "DB388"'),
    (49,13,"learning",'Watch: "Go for Magnificence" — Amway App / amway.com > Education'),
    (49,13,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (49,13,"networking","Learn about Amway Dish products — tell 1+ person about them"),
    (49,13,"habit",  "Clear Topics Daily"),
    (49,13,"habit",  "Check DreamStream Daily"),
    (49,13,"habit",  "Send Topic to coach with takeaways and a question"),

    (50,13,"learning",'Listen: "SL 2024 — Bill and Sandy Hawkins" — WWG App > WWG Store > Search "1454"'),
    (50,13,"learning",'Watch: "Go, Team" — Amway App / amway.com > Education'),
    (50,13,"learning",'Learn Amway Home product demos — Search "Amway Home Demos" in Amway App > Education > Resources tab'),
    (50,13,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (50,13,"event",  "Complete a virtual or live Amway Home Retail Event with 1+ person"),
    (50,13,"habit",  "Clear Topics Daily"),
    (50,13,"habit",  "Check DreamStream Daily"),
    (50,13,"habit",  "Send Topic to coach with takeaways and a question"),

    (51,13,"learning",'Listen: "FED 2022 — Herschelman and Eaton" — WWG App > WWG Store > Search "1335"'),
    (51,13,"learning",'Watch: "Overcoming Great Adversity" — WWG App > WWG Store > Search "1487"'),
    (51,13,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (51,13,"networking","Learn about Nutrilite Ashwagandha — tell 1+ person about it"),
    (51,13,"networking","Follow up with Amway Home Event customers — check if they received their orders"),
    (51,13,"habit",  "Clear Topics Daily"),
    (51,13,"habit",  "Check DreamStream Daily"),
    (51,13,"habit",  "Send Topic to coach with takeaways and a question"),

    (52,13,"learning",'Listen: "Hope is Your Super Power" — WWG App > WWG Store > Search "1170"'),
    (52,13,"learning",'Watch: "Lead with Influence" — Amway App / amway.com > Education'),
    (52,13,"networking","Practice Interjections: 10 Make Awares + 4 Drop Messages this week"),
    (52,13,"networking","Follow up with Amway Home Event customers — check if they are enjoying their orders"),
    (52,13,"habit",  "Clear Topics Daily"),
    (52,13,"habit",  "Check DreamStream Daily"),
    (52,13,"habit",  "Send Topic to coach with takeaways and a question"),
    (52,13,"event",  "GREAT JOB! You finished the 52-Week IBO Training. Celebrate your growth — then do it again!"),
]

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=11, color="000000", italic=False):
    return Font(bold=True, size=size, color=color, italic=italic)

def thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

CAT_COLORS = {
    "learning":    ("EEF2FF", "3730A3"),   # indigo light / dark
    "habit":       ("F0FDFA", "0F766E"),   # teal light / dark
    "networking":  ("FFFBEB", "92400E"),   # amber light / dark
    "event":       ("FDF2F8", "9D174D"),   # pink light / dark
}

CAT_LABELS = {
    "learning": "Learning",
    "habit":    "Daily Habit",
    "networking": "Networking",
    "event":    "Event",
}

WEEK_ROW_FILLS = ["FFFFFF", "F8FAFC"]   # alternating by week number

# ── BUILD WORKBOOK ────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default Sheet

# ─── SHEET 1: All Tasks ───────────────────────────────────────────────────────
ws_tasks = wb.create_sheet("All Tasks")
ws_tasks.sheet_properties.tabColor = "22C55E"  # green

# Headers
headers = ["Week", "Month", "Category", "Task", "Done?"]
col_widths = [8, 8, 14, 80, 8]
header_fill = fill("1E293B")
header_font = bold_font(11, "FFFFFF")

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws_tasks.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_tasks.column_dimensions[get_column_letter(col)].width = w

ws_tasks.row_dimensions[1].height = 22
ws_tasks.freeze_panes = "A2"

# Data rows
dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
ws_tasks.add_data_validation(dv)

row = 2
for (week, month, cat, task) in WEEKS:
    bg = WEEK_ROW_FILLS[(week - 1) % 2]
    cat_bg, cat_fg = CAT_COLORS[cat]

    ws_tasks.cell(row=row, column=1, value=week).alignment = Alignment(horizontal="center", vertical="top")
    ws_tasks.cell(row=row, column=2, value=month).alignment = Alignment(horizontal="center", vertical="top")

    cat_cell = ws_tasks.cell(row=row, column=3, value=CAT_LABELS[cat])
    cat_cell.fill = fill(cat_bg)
    cat_cell.font = Font(bold=True, size=10, color=cat_fg)
    cat_cell.alignment = Alignment(horizontal="center", vertical="top")

    task_cell = ws_tasks.cell(row=row, column=4, value=task)
    task_cell.alignment = Alignment(wrap_text=True, vertical="top")

    done_cell = ws_tasks.cell(row=row, column=5, value=False)
    done_cell.alignment = Alignment(horizontal="center", vertical="top")
    dv.add(done_cell)

    # Row background (columns 1,2,4,5)
    for col in [1, 2, 4, 5]:
        ws_tasks.cell(row=row, column=col).fill = fill(bg)

    ws_tasks.row_dimensions[row].height = 30
    row += 1

# Conditional formatting: checked rows green
total_rows = row - 1
green_fill = PatternFill("solid", fgColor="DCFCE7")
strike_font = Font(strikethrough=True, color="6B7280")
ws_tasks.conditional_formatting.add(
    f"A2:E{total_rows}",
    FormulaRule(formula=["$E2=TRUE"], fill=green_fill)
)

# Auto-filter
ws_tasks.auto_filter.ref = f"A1:E{total_rows}"

# ─── SHEET 2: Monthly Goals ───────────────────────────────────────────────────
ws_goals = wb.create_sheet("Monthly Goals")
ws_goals.sheet_properties.tabColor = "A855F7"  # purple

ws_goals.column_dimensions["A"].width = 10
ws_goals.column_dimensions["B"].width = 72
ws_goals.column_dimensions["C"].width = 10

# Header row
for col, h in enumerate(["Month", "Goal", "Done?"], 1):
    c = ws_goals.cell(row=1, column=col, value=h)
    c.fill = fill("1E293B")
    c.font = bold_font(11, "FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
ws_goals.row_dimensions[1].height = 22
ws_goals.freeze_panes = "A2"

dv_goals = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
ws_goals.add_data_validation(dv_goals)

MONTH_FILLS = [
    "EEF2FF","F0FDF4","FFF7ED","FDF2F8","F0FDFA","FEFCE8",
    "EFF6FF","FFF1F2","F5F3FF","ECFDF5","FEF9C3","F0F9FF","FFF5F5"
]

row = 2
for month in range(1, 14):
    goals = MONTHLY_GOALS[month]
    mfill = fill(MONTH_FILLS[month - 1])

    # Month header sub-row
    hdr = ws_goals.cell(row=row, column=1, value=f"Month {month}")
    hdr.fill = fill("334155")
    hdr.font = bold_font(10, "FFFFFF")
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    ws_goals.cell(row=row, column=2).fill = fill("334155")
    ws_goals.cell(row=row, column=3).fill = fill("334155")
    ws_goals.row_dimensions[row].height = 18
    row += 1

    for g in goals:
        ws_goals.cell(row=row, column=1, value=month).fill = mfill
        ws_goals.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")

        gc = ws_goals.cell(row=row, column=2, value=g)
        gc.fill = mfill
        gc.alignment = Alignment(wrap_text=True, vertical="top")

        dc = ws_goals.cell(row=row, column=3, value=False)
        dc.fill = mfill
        dc.alignment = Alignment(horizontal="center", vertical="top")
        dv_goals.add(dc)

        ws_goals.row_dimensions[row].height = 28
        row += 1

goals_total = row - 1
ws_goals.conditional_formatting.add(
    f"A2:C{goals_total}",
    FormulaRule(formula=["$C2=TRUE"], fill=PatternFill("solid", fgColor="DCFCE7"))
)
ws_goals.auto_filter.ref = f"A1:C1"

# ─── SHEET 3: Dashboard ──────────────────────────────────────────────────────
ws_dash = wb.create_sheet("Dashboard", 0)   # first tab
ws_dash.sheet_properties.tabColor = "3B82F6"  # blue
ws_dash.column_dimensions["A"].width = 10
ws_dash.column_dimensions["B"].width = 12
ws_dash.column_dimensions["C"].width = 14
ws_dash.column_dimensions["D"].width = 14
ws_dash.column_dimensions["E"].width = 12
ws_dash.column_dimensions["F"].width = 20

# Title
title = ws_dash.cell(row=1, column=1, value="52-Week IBO Tracker — Jackie's Training Journey")
title.font = bold_font(16, "1E293B")
ws_dash.merge_cells("A1:F1")
ws_dash.row_dimensions[1].height = 30

ws_dash.cell(row=2, column=1, value="Track your progress week by week. Update checkboxes in 'All Tasks' and 'Monthly Goals' sheets.")
ws_dash.cell(row=2, column=1).font = Font(italic=True, color="64748B", size=10)
ws_dash.merge_cells("A2:F2")

# Summary stats
ws_dash.row_dimensions[4].height = 20
total_tasks = len(WEEKS)
for col, (label, val_formula) in enumerate([
    ("Total Tasks", str(total_tasks)),
    ("Done", f"=COUNTIF('All Tasks'!E:E,TRUE)"),
    ("Remaining", f"={total_tasks}-COUNTIF('All Tasks'!E:E,TRUE)"),
    ("% Complete", f"=COUNTIF('All Tasks'!E:E,TRUE)/{total_tasks}"),
], 1):
    lc = ws_dash.cell(row=4, column=col, value=label)
    lc.fill = fill("1E293B")
    lc.font = bold_font(10, "FFFFFF")
    lc.alignment = Alignment(horizontal="center")
    ws_dash.row_dimensions[4].height = 20

    vc = ws_dash.cell(row=5, column=col, value=val_formula if val_formula.startswith("=") else int(val_formula))
    vc.fill = fill("EFF6FF")
    vc.font = bold_font(14, "1E293B")
    vc.alignment = Alignment(horizontal="center")
    ws_dash.row_dimensions[5].height = 32

ws_dash.cell(row=5, column=4).number_format = "0%"

# Per-week progress table
ws_dash.row_dimensions[8].height = 20
for col, h in enumerate(["Week", "Month", "Tasks Done", "Total Tasks", "% Done", "Status"], 1):
    c = ws_dash.cell(row=8, column=col, value=h)
    c.fill = fill("334155")
    c.font = bold_font(10, "FFFFFF")
    c.alignment = Alignment(horizontal="center")

# Build per-week summary
from collections import defaultdict
week_totals = defaultdict(int)
week_months = {}
for (wk, mo, cat, task) in WEEKS:
    week_totals[wk] += 1
    week_months[wk] = mo

# Map rows: tasks sheet row range per week
week_row_start = {}
week_row_end = {}
cur_week = None
r = 2
for (wk, mo, cat, task) in WEEKS:
    if wk != cur_week:
        week_row_start[wk] = r
        if cur_week is not None:
            week_row_end[cur_week] = r - 1
        cur_week = wk
    r += 1
week_row_end[52] = r - 1

for i, wk in enumerate(range(1, 53)):
    row_idx = 9 + i
    mo = week_months[wk]
    total = week_totals[wk]
    r_start = week_row_start[wk]
    r_end = week_row_end[wk]

    ws_dash.cell(row=row_idx, column=1, value=wk).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=row_idx, column=2, value=mo).alignment = Alignment(horizontal="center")
    done_f = f"=COUNTIFS('All Tasks'!A:A,{wk},'All Tasks'!E:E,TRUE)"
    ws_dash.cell(row=row_idx, column=3, value=done_f).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=row_idx, column=4, value=total).alignment = Alignment(horizontal="center")
    pct_f = f"=C{row_idx}/{total}"
    pct_c = ws_dash.cell(row=row_idx, column=5, value=pct_f)
    pct_c.number_format = "0%"
    pct_c.alignment = Alignment(horizontal="center")
    status_f = f'=IF(C{row_idx}={total},"Complete ✓",IF(C{row_idx}>0,"In Progress","Not Started"))'
    ws_dash.cell(row=row_idx, column=6, value=status_f).alignment = Alignment(horizontal="center")

    bg = WEEK_ROW_FILLS[(wk - 1) % 2]
    for col in range(1, 7):
        ws_dash.cell(row=row_idx, column=col).fill = fill(bg)

# Conditional formatting on Dashboard status column
ws_dash.conditional_formatting.add(
    f"F9:F{9+51}",
    FormulaRule(formula=[f"=$F9=\"Complete ✓\""], fill=PatternFill("solid", fgColor="DCFCE7"))
)

# ─── SHEET 4: Progress Chart ─────────────────────────────────────────────────
ws_chart = wb.create_sheet("Progress Chart")
ws_chart.sheet_properties.tabColor = "F97316"  # orange

ws_chart.cell(row=1, column=1, value="Weekly Progress Chart").font = bold_font(14, "1E293B")
ws_chart.merge_cells("A1:H1")

chart = BarChart()
chart.type = "col"
chart.title = "Tasks Completed per Week"
chart.y_axis.title = "Tasks Done"
chart.x_axis.title = "Week"
chart.style = 10
chart.height = 14
chart.width = 28

data_ref = Reference(ws_dash, min_col=3, min_row=8, max_row=60)
cats_ref = Reference(ws_dash, min_col=1, min_row=9, max_row=60)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.shape = 4

ws_chart.add_chart(chart, "A3")

# ─── SAVE ─────────────────────────────────────────────────────────────────────

out_path = os.path.join(os.path.dirname(__file__), "tracker.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Total task rows: {total_tasks}")
